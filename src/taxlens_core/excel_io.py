from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .constants import (
    ASSET_COLUMNS,
    ASSET_LEDGER_SHEET,
    MAX_FILE_SIZE_BYTES,
    MAX_RECORDS,
    MIN_RECORDS,
)
from .models import IssueSeverity, ValidationIssue


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def read_and_validate_workbook(path: str | Path) -> tuple[list[dict[str, Any]], list[ValidationIssue]]:
    file_path = Path(path)
    issues: list[ValidationIssue] = []

    if file_path.suffix.lower() != ".xlsx":
        issues.append(ValidationIssue(
            "INVALID_FILE_TYPE", "仅支持 .xlsx 文件", IssueSeverity.ERROR
        ))
        return [], issues
    if not file_path.exists():
        issues.append(ValidationIssue(
            "FILE_NOT_FOUND", "文件不存在", IssueSeverity.ERROR
        ))
        return [], issues
    if file_path.stat().st_size > MAX_FILE_SIZE_BYTES:
        issues.append(ValidationIssue(
            "FILE_TOO_LARGE", "文件大小超过 2 MB", IssueSeverity.ERROR
        ))
        return [], issues

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        issues.append(ValidationIssue(
            "WORKBOOK_UNREADABLE", f"Excel 文件无法读取：{exc}", IssueSeverity.ERROR
        ))
        return [], issues

    try:
        if ASSET_LEDGER_SHEET not in workbook.sheetnames:
            issues.append(ValidationIssue(
                "MISSING_REQUIRED_SHEET",
                f"缺少必需工作表：{ASSET_LEDGER_SHEET}",
                IssueSeverity.ERROR,
            ))
            return [], issues

        sheet = workbook[ASSET_LEDGER_SHEET]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        headers = [str(value).strip() if value is not None else "" for value in first_row]
        missing = [column for column in ASSET_COLUMNS if column not in headers]
        duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
        if missing:
            issues.append(ValidationIssue(
                "MISSING_REQUIRED_COLUMNS",
                "缺少必需字段：" + ", ".join(missing),
                IssueSeverity.ERROR,
            ))
        if duplicate_headers:
            issues.append(ValidationIssue(
                "DUPLICATE_COLUMNS",
                "存在重复字段：" + ", ".join(duplicate_headers),
                IssueSeverity.ERROR,
            ))
        if missing or duplicate_headers:
            return [], issues

        header_index = {name: index for index, name in enumerate(headers)}
        records: list[dict[str, Any]] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if _is_blank_row(values):
                continue
            records.append({
                column: values[header_index[column]] if header_index[column] < len(values) else None
                for column in ASSET_COLUMNS
            } | {"_row_number": row_number})

        if not (MIN_RECORDS <= len(records) <= MAX_RECORDS):
            issues.append(ValidationIssue(
                "INVALID_RECORD_COUNT",
                f"资产记录数必须为 {MIN_RECORDS}–{MAX_RECORDS} 条，当前为 {len(records)} 条",
                IssueSeverity.ERROR,
            ))

        ids: dict[str, list[int]] = {}
        for record in records:
            raw_id = record.get("asset_id")
            asset_id = str(raw_id).strip() if raw_id is not None else ""
            if asset_id:
                ids.setdefault(asset_id, []).append(record["_row_number"])
        for asset_id, rows in ids.items():
            if len(rows) > 1:
                issues.append(ValidationIssue(
                    "DUPLICATE_ASSET_ID",
                    f"asset_id {asset_id} 重复，行号：{', '.join(map(str, rows))}",
                    IssueSeverity.ERROR,
                    field="asset_id",
                ))
        return records, issues
    finally:
        workbook.close()

